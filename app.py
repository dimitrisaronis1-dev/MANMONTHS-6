import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
import copy # Import standard copy module
from datetime import datetime
from dateutil.relativedelta import relativedelta
import re
import streamlit as st # Import streamlit
import os # Import os module to check for file existence
import requests # Import requests to download files from URLs
import io # Import io to handle byte streams
import random # Import random module

# Initialize Streamlit app
st.set_page_config(layout='wide')

# --- Template and Logo URLs ---
TEMPLATE_URL = "https://raw.githubusercontent.com/dimitrisaronis1-dev/MANMONTHS-6/main/AM%20TEST%201.xlsx"
LOGO_URL = "https://raw.githubusercontent.com/dimitrisaronis1-dev/MANMONTHS-6/main/SPACE%20LOGO_colored%20horizontal.png"

# --- Display Logo and Title ---
col1, col2 = st.columns([3, 1]) # Adjust column ratio as needed
with col1:
    st.write("### Κατανομή Ανθρωπομηνών (Person-Months Allocation)")
with col2:
    st.image(LOGO_URL, width=378) # 10cm is approximately 378 pixels

# Replaced original colab file upload with streamlit file uploader
input_file = st.file_uploader("👉 Ανέβασε το INPUT excel (μόνο 2 στήλες)", type=["xlsx"])

yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Define thin border style
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Wrap the main script logic within an if input_file is not None: block
if input_file is not None:
    # Download the template file
    try:
        response = requests.get(TEMPLATE_URL)
        response.raise_for_status() # Raise an exception for HTTP errors
        template_buffer = io.BytesIO(response.content)
    except requests.exceptions.RequestException as e:
        st.error(f"Error downloading template file: {e}")
        st.stop()

    # ------------------------------------------------
    # Συναρτήσεις ημερομηνιών
    # ------------------------------------------------
    def parse_date(text, is_start=True):
        text = str(text).strip() # Ensure text is always a string and stripped

        # Handle 'σήμερα' (today) keyword
        if "σήμερα" in text.lower() or "simera" in text.lower(): # Added 'simera' as a fallback, user provided 'σήμερα'
            if not is_start: # 'Σήμερα' is treated as an end date
                return datetime.today()
            else:
                # If 'σήμερα' is explicitly used as a start date, we could either raise an error
                # or treat it as today. Based on the example, it's an end date. For simplicity,
                # if it's a start date and contains 'σήμερα', we'll let it pass to the regex checks
                # which will likely fail, implying it's not a supported start date format in combination.
                pass # Let it fall through to other parsing if it's a start date and contains 'σήμερα'

        if re.match(r"^\d{4}$", text): # Check for YYYY format
            if is_start:
                return datetime.strptime("01/01/" + text, "%d/%m/%Y") # Start of the year
            else:
                return datetime.strptime("31/12/" + text, "%d/%m/%Y") # End of the year
        elif re.match(r"^\d{1,2}/\d{4}$", text): # Check for MM/YYYY format (updated regex)
            if is_start:
                # For MM/YYYY as start, it's always the 1st of the month
                return datetime.strptime("01/" + text, "%d/%m/%Y")
            else:
                # For MM/YYYY as end, it's the last day of the month
                d = datetime.strptime("01/" + text, "%d/%m/%Y")
                return d + relativedelta(months=1) - relativedelta(days=1)
        elif re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", text): # Check for DD/MM/YYYY format
            return datetime.strptime(text, "%d/%m/%Y")
        else:
            # If none of the above formats match, raise a more informative error
            raise ValueError(f"Unsupported date format: '{text}'. Expected 'YYYY', 'M/YYYY' or 'MM/YYYY', 'D/M/YYYY' or 'DD/MM/YYYY', or 'Σήμερα' (for end date).")

    def parse_period(p):
        p_cleaned = str(p).strip().replace("—", "-").replace("–", "-") # Ensure p is a string and cleaned

        # Check if the cleaned string is a single year (e.g., "2022")
        if re.match(r"^\d{4}$", p_cleaned):
            # If it's a single year, treat it as the entire year
            return parse_date(p_cleaned, True), parse_date(p_cleaned, False)

        parts = p_cleaned.split("-")
        if len(parts) != 2:
            raise ValueError(f"Invalid period format: '{p}'. Expected 'YYYY' or 'START_DATE-END_DATE'.")
        a, b = parts
        return parse_date(a, True), parse_date(b, False)

    def month_range(start, end):
        current = datetime(start.year, start.month, 1)
        end = datetime(end.year, end.month, 1)
        out = []
        while current <= end:
            out.append((current.year, current.month))
            current += relativedelta(months=1)
        return out

    # Function to determine if a color is light or dark (for text readability)
    def is_light_color(hex_color):
        hex_color = hex_color.lstrip('#')
        rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        # Calculate luminance (Y = 0.299R + 0.587G + 0.114B)
        luminance = (0.299 * rgb[0] + 0.587 * rgb[1] + 0.114 * rgb[2]) / 255
        return luminance > 0.5 # Threshold can be adjusted

    # ------------------------------------------------
    # Διαβάζουμε INPUT και βρίσκουμε στήλες
    # ------------------------------------------------
    # openpyxl can directly handle UploadedFile objects
    wb_in = openpyxl.load_workbook(input_file)
    ws_in = wb_in.active

    headers = {}
    for c in range(1, ws_in.max_column + 1):
        val = str(ws_in.cell(1,c).value).strip()
        headers[val] = c

    if "ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ" not in headers or "ΑΝΘΡΩΠΟΜΗΝΕΣ" not in headers:
        st.error("Το input πρέπει να έχει στήλες: ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ και ΑΝΘΡΩΠΟΜΗΝΕΣ")
        st.stop() # Stop execution if headers are missing

    PERIOD_COL = headers["ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ"]
    AM_COL = headers["ΑΝΘΡΩΠΟΜΗΝΕΣ"]

    data = []
    all_months = set()
    project_counter = 0 # Initialize project counter

    for r in range(2, ws_in.max_row + 1):
        period = ws_in.cell(r, PERIOD_COL).value
        am_raw = ws_in.cell(r, AM_COL).value
        try:
            am = int(am_raw) if am_raw is not None else 0
        except (ValueError, TypeError):
            am = 0 # Handle non-numeric or None values

        if not period or am == 0: # Skip rows with no period or 0 AMs
            continue

        try:
            start, end = parse_period(str(period))
        except ValueError as e:
            st.warning(f"Skipping row {r} due to period parsing error: {e}") # Convert print to st.warning
            continue

        months = month_range(start, end)
        months_in_period_count = len(months)

        if months_in_period_count > 0: # Avoid division by zero
            am_per_month_ratio = am / months_in_period_count
        else:
            am_per_month_ratio = 0

        data.append({
            "project_id": project_counter, # Assign a unique ID
            "period_str": period,
            "original_am": am,
            "months_in_period": months,
            "months_in_period_count": months_in_period_count,
            "am_per_month_ratio": am_per_month_ratio,
            "allocated_am": 0,
            "unallocated_am": am
        })
        project_counter += 1 # Increment for the next project

        for m in months:
            all_months.add(m)

    all_months = sorted(all_months)
    years = sorted(set(y for y,m in all_months))

    # Project Prioritization: Sort data before allocation
    # Prioritize projects with shorter durations (fewer months in period)
    data.sort(key=lambda x: x["months_in_period_count"])

    # ------------------------------------------------
    # Ανοίγουμε TEMPLATE
    # ------------------------------------------------
    # Use the downloaded template buffer
    wb = openpyxl.load_workbook(template_buffer)
    ws = wb.active

    # Freeze the first 3 columns (A, B, C) - this means the freeze point is at D1
    ws.freeze_panes = 'D1'

    # Adjusted START_ROW, YEAR_ROW, MONTH_ROW
    START_ROW_DATA = 4 # Data starts here
    YEAR_ROW = 2  # Years go here
    MONTH_ROW = 3 # Months go here
    YEARLY_TOTAL_ROW = START_ROW_DATA + 1 # New row for yearly totals
    START_COL = 5 # Month 1 of first year starts in column E
    MAX_YEARLY_CAPACITY = 11 # Define MAX_YEARLY_CAPACITY here

    # ------------------------------------------------
    # Καθαρισμός παλιάς περιοχής
    # ------------------------------------------------
    # Unmerge any previous cells in relevant header and data rows to prevent conflicts
    merged_cells_to_unmerge = []
    for cell_range_str in list(ws.merged_cells.ranges): # Iterate over a copy of the list
        min_col_mc, min_row_mc, max_col_mc, max_row_mc = openpyxl.utils.cell.range_boundaries(str(cell_range_str))
        if (
           (min_row_mc <= YEAR_ROW <= max_row_mc) or \
           (min_row_mc <= MONTH_ROW <= max_col_mc) or \
           (min_row_mc <= START_ROW_DATA <= max_row_mc) or \
           (min_row_mc <= YEARLY_TOTAL_ROW <= max_row_mc)
        ):
            merged_cells_to_unmerge.append(cell_range_str)

    for cell_range_str in merged_cells_to_unmerge:
        ws.unmerge_cells(str(cell_range_str))

    # Determine the maximum column that could have been written to, to ensure full clearing
    max_col_to_clear = max(START_COL + len(years) * 12, ws.max_column + 1)

    # Clear specific rows: Year row (2), MONTH_ROW (3), START_ROW_DATA (4), and YEARLY_TOTAL_ROW (5)
    rows_to_clear_completely = [YEAR_ROW, MONTH_ROW, START_ROW_DATA, YEARLY_TOTAL_ROW]

    for r_clear in rows_to_clear_completely:
        for c_clear in range(1, max_col_to_clear):
            ws.cell(r_clear, c_clear).value = None
            ws.cell(r_clear, c_clear).fill = PatternFill() # Also clear fill in these rows

    # Clear all data cells (values and fills) starting from START_ROW_DATA + 2 (row 6 onwards)
    # This accounts for the new total row at START_ROW_DATA + 1
    for r_clear in range(START_ROW_DATA + 2, ws.max_row + 1):
        for c_clear in range(1, max_col_to_clear):
            ws.cell(r_clear,c_clear).value = None
            ws.cell(r_clear,c_clear).fill = PatternFill() # Clear fill as well

    # ------------------------------------------------
    # Initialize Allocation Data Structures
    # ------------------------------------------------
    yearly_am_totals = {year: 0 for year in years}
    # Tracks (year, month) to project_id to allow reporting which project filled a slot
    month_allocation_status = {(y, m): None for y in years for m in range(1, 13)}

    # ------------------------------------------------
    # Χτίσιμο ετών & μηνών
    # ------------------------------------------------
    col = START_COL
    month_col_map = {}

    for y in years:
        year_start_col = col # Store the starting column for this year
        year_header_cell = ws.cell(YEAR_ROW, col) # Get the cell where year will be displayed

        # Apply random fill to the merged year cell
        # Directly call random.randint to avoid NameError with lambda in some Streamlit contexts
        random_color_hex = '%02X%02X%02X' % (random.randint(0,255), random.randint(0,255), random.randint(0,255))
        year_header_cell.fill = PatternFill(start_color=random_color_hex, end_color=random_color_hex, fill_type="solid")

        # Set font color for the year based on its background color
        if not is_light_color(random_color_hex):
            year_header_cell.font = Font(color="FFFFFF") # White text for dark background
        else:
            year_header_cell.font = Font(color="000000") # Black text for light background

        for m in range(1,13):
            ws.cell(MONTH_ROW, col).value = m
            month_col_map[(y,m)] = col
            col += 1
        year_end_col = col - 1 # The last column for this year (12 months)

        # Merge cells for the current year
        ws.merge_cells(start_row=YEAR_ROW, start_column=year_start_col, end_row=YEAR_ROW, end_column=year_end_col)
        year_header_cell.value = y # Year will be written in the first cell of the merged block

    # Apply borders to year and month headers
    for c_border in range(START_COL, col):
        ws.cell(YEAR_ROW, c_border).border = thin_border
        ws.cell(MONTH_ROW, c_border).border = thin_border

    # Add headers for the yearly total row
    ws.cell(YEARLY_TOTAL_ROW, 2).value = "ΕΤΗΣΙΑ ΣΥΝΟΛΑ"
    ws.cell(YEARLY_TOTAL_ROW, 2).font = Font(bold=True)
    ws.cell(YEARLY_TOTAL_ROW, 2).border = thin_border

    # ------------------------------------------------
    # Γραμμές & μπάρες (Greedy Allocation)
    # ------------------------------------------------
    row = START_ROW_DATA + 2 # Data starts from the new START_ROW_DATA + 2 (after yearly totals row)

    unallocated_projects = []
    yearly_overages = {}

    for project_idx, project_data in enumerate(data):
        period_str = project_data["period_str"]
        original_am = project_data["original_am"]
        months_in_period = project_data["months_in_period"]
        project_id = project_data["project_id"]
        allocated_count = 0
        unallocated_count = original_am
        project_unallocated_reason = []

        # Apply borders to columns B and C for current data row
        ws.cell(row,2).value = period_str
        ws.cell(row,2).border = thin_border
        ws.cell(row,3).value = original_am
        ws.cell(row,3).border = thin_border

        # Iterate through months in chronological order for greedy allocation
        for (y, m) in sorted(months_in_period): # Ensure chronological processing
            if allocated_count >= original_am: # All AMs for this project are allocated
                break

            if (y,m) in month_col_map:
                # Check if year is already at capacity
                if yearly_am_totals[y] >= MAX_YEARLY_CAPACITY:
                    if f"Year {y} capacity reached" not in project_unallocated_reason:
                        project_unallocated_reason.append(f"Year {y} capacity reached")
                    continue # Skip this month, cannot allocate due to yearly cap

                # Check if month slot is already taken by another project
                if month_allocation_status[(y,m)] is not None:
                    # Record which project took the slot
                    occupying_project_id = month_allocation_status[(y,m)]
                    if f"Month {m}/{y} already allocated by Project {occupying_project_id}" not in project_unallocated_reason:
                        project_unallocated_reason.append(f"Month {m}/{y} already allocated by Project {occupying_project_id}")
                    continue # Skip this month, slot is taken

                # Allocate one person-month to this slot
                cell_to_fill = ws.cell(row, month_col_map[(y,m)])
                cell_to_fill.value = 'X' # Mark allocation
                cell_to_fill.fill = yellow
                cell_to_fill.border = thin_border

                yearly_am_totals[y] += 1
                month_allocation_status[(y,m)] = project_id # Mark month as allocated by this project ID
                allocated_count += 1
                unallocated_count -= 1

        project_data["allocated_am"] = allocated_count
        project_data["unallocated_am"] = unallocated_count

        if unallocated_count > 0:
            unallocated_projects.append({
                "period": period_str,
                "original_am": original_am,
                "allocated_am": allocated_count,
                "unallocated_am": unallocated_count,
                "reasons": "; ".join(list(set(project_unallocated_reason))) # Use set to remove duplicate reasons
            })
            # Highlight the original AM if not fully allocated (e.g., red text)
            ws.cell(row, 3).font = Font(color="FF0000", bold=True)
        else:
            ws.cell(row, 3).font = Font(color="000000")

        # Apply borders to all month columns for the current data row (even if no allocation happened)
        for c_border in range(START_COL, col):
            ws.cell(row, c_border).border = thin_border

        row += 1

    # Populate and style the YEARLY_TOTAL_ROW
    for y in years:
        if y in yearly_am_totals:
            col_for_year_total = month_col_map[(y, 1)] # Get first month's column for the year
            # Handle merged cells for yearly total row
            year_month_cols = [month_col_map[(y, m)] for m in range(1, 13) if (y, m) in month_col_map]
            if year_month_cols:
                year_start_col = min(year_month_cols)
                year_end_col = max(year_month_cols)
                if year_start_col != year_end_col: # Only merge if there's more than one column for the year
                    ws.merge_cells(start_row=YEARLY_TOTAL_ROW, start_column=year_start_col, end_row=YEARLY_TOTAL_ROW, end_column=year_end_col)
                col_for_year_total = year_start_col
            else:
                continue # Skip if no month columns found for the year

            total_cell = ws.cell(YEARLY_TOTAL_ROW, col_for_year_total)
            total_cell.value = yearly_am_totals[y]
            total_cell.font = Font(bold=True)
            total_cell.border = thin_border
            total_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Highlight if yearly total meets or exceeds capacity
            if yearly_am_totals[y] >= MAX_YEARLY_CAPACITY:
                # Apply to the merged year header cell (first column of the year)
                year_header_cell = ws.cell(YEAR_ROW, col_for_year_total)
                year_header_cell.fill = red_fill
                year_header_cell.font = Font(color="FFFFFF", bold=True)

                total_cell.fill = red_fill
                total_cell.font = Font(color="FFFFFF", bold=True)
                if yearly_am_totals[y] > MAX_YEARLY_CAPACITY:
                    yearly_overages[y] = yearly_am_totals[y] - MAX_YEARLY_CAPACITY
            elif yearly_am_totals[y] > 0:
                total_cell.fill = green_fill # Green for under capacity but allocated
                total_cell.font = Font(color="000000", bold=True)

    # Set fixed width for month columns after all content has been added
    for c_width in range(START_COL, col):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_width)].width = 2.5

    # ------------------------------------------------
    # Output Summary
    # ------------------------------------------------
    st.write("\n--- Allocation Summary ---") # Convert print to st.write
    st.write(f"Max yearly capacity per year: {MAX_YEARLY_CAPACITY} person-months") # Convert print to st.write

    st.write("\nYearly Person-Month Totals:") # Convert print to st.write
    for year, total_am in sorted(yearly_am_totals.items()):
        status = "(Capacity Reached)" if total_am >= MAX_YEARLY_CAPACITY else ""
        if year in yearly_overages: # Check if there was an actual overage
            status = f"(OVER CAPACITY by {yearly_overages[year]})"
        st.write(f"  Year {year}: {total_am} {status}") # Convert print to st.write


    if unallocated_projects:
        st.write("\nProjects with Unallocated Person-Months:") # Convert print to st.write
        for proj in unallocated_projects:
            st.write(f"  Period: {proj['period']}, Original AM: {proj['original_am']}, Allocated AM: {proj['allocated_am']}, Unallocated AM: {proj['unallocated_am']}") # Convert print to st.write
            if proj['reasons']:
                st.write(f"    Reasons for unallocation: {proj['reasons']}") # Convert print to st.write
    else:
        st.success("\nAll person-months were allocated successfully.") # Convert print to st.success


    # ------------------------------------------------
    # Save & download
    # ------------------------------------------------
    # Extract base name without extension
    base_name = os.path.splitext(input_file.name)[0] # Use input_file.name for Streamlit UploadedFile
    # Construct new output filename
    output = f"{base_name}_ΚΑΤΑΝΟΜΗ ΑΜ.xlsx"

    # Rename the active worksheet to 'ΑΝΑΛΥΣΗ'
    ws.title = 'ΑΝΑΛΥΣΗ'

    # Create a new sheet named 'CV' and copy content from ws_in.active
    cv_sheet = wb.create_sheet(title='CV', index=0) # Add as the first sheet
    for row_idx, row_data in enumerate(ws_in.iter_rows()): # Renamed 'row' to 'row_data' to avoid conflict with outer 'row'
        for col_idx, cell in enumerate(row_data):
            new_cell = cv_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell.value)
            # Copy styles (fills, fonts, borders, number format, etc.)
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.border = copy.copy(cell.border)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.number_format = cell.number_format

    # Copy column dimensions (widths)
    for col_idx in range(1, ws_in.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if col_letter in ws_in.column_dimensions:
            cv_sheet.column_dimensions[col_letter].width = ws_in.column_dimensions[col_letter].width


    # Add the formula to cell A6 of the 'ΑΝΑΛΥΣΗ' sheet
    ws['A6'] = '=MATCH(B6,CV!$B$2:$B$100,0)'

    # Save to a temporary buffer or file for download
    import io
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0) # Rewind the buffer to the beginning

    # Streamlit download button
    st.download_button(
        label="Download Processed Excel file",
        data=output_buffer,
        file_name=output,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.info("Παρακαλώ ανεβάστε ένα αρχείο Excel για να ξεκινήσετε την επεξεργασία.")